import json
import csv
import os
from io import StringIO
from django.contrib import admin, messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db import transaction
from django.core.files.storage import default_storage
import openpyxl
import logging
from .base import BaseAdmin
from ..models import Product, Category
from ..forms import ProductImportForm, ProductExportForm

logger = logging.getLogger(__name__)

@admin.register(Product)
class ProductAdmin(BaseAdmin):
    list_display = ('id', 'name', 'category', 'price', 'created_at', 'is_active')
    search_fields = ('name', 'description')
    list_filter = ('category', 'is_active')
    actions = BaseAdmin.actions + ['import_products', 'export_products']

    def save_model(self, request, obj, form, change):
        if change:
            logger.info(f'Товар изменён: {obj}')
        else:
            logger.info(f'Создан новый товар: {obj}')
        super().save_model(request, obj, form, change)

    def delete_model(self, request, obj):
        logger.info(f'Товар мягко удалён: {obj}')
        obj.soft_delete()

    def soft_delete_selected(self, request, queryset):
        for obj in queryset:
            obj.soft_delete()
        self.message_user(request, "Выбранные товары мягко удалены.")
    soft_delete_selected.short_description = "Мягко удалить выбранные товары"

    def hard_delete_selected(self, request, queryset):
        for obj in queryset:
            logger.info(f'Товар полностью удалён: {obj}')
            obj.delete()
        self.message_user(request, "Выбранные товары полностью удалены.")
    hard_delete_selected.short_description = "Полностью удалить выбранные товары"

    def get_queryset(self, request):
        return Product.objects.all()

    def import_products(self, request):
        logger.info("Начало импорта товаров")

        # Если это POST-запрос от формы импорта
        if 'import_form_submit' in request.POST:
            logger.info("Получен POST-запрос от формы импорта")
            form = ProductImportForm(request.POST, request.FILES)
            if form.is_valid():
                logger.info("Форма валидна")
                file = request.FILES['file']
                file_format = form.cleaned_data['file_format']
                errors = []
                imported_count = 0

                try:
                    with transaction.atomic():
                        if file_format == 'json':
                            data = self.parse_json(file)
                        elif file_format == 'csv':
                            data = self.csv_to_json(file)
                        elif file_format == 'xlsx':
                            data = self.excel_to_json(file)
                        else:
                            logger.error("Неподдерживаемый формат файла")
                            messages.error(request, "Неподдерживаемый формат файла.")
                            return redirect('admin:shop_product_changelist')

                        logger.info(f"Обработка {len(data)} записей из файла")
                        for index, row in enumerate(data):
                            try:
                                if not row.get('name'):
                                    errors.append(f"Запись {index + 1}: Поле 'name' обязательно.")
                                    continue
                                if 'price' not in row or row['price'] is None:
                                    errors.append(f"Запись {index + 1}: Поле 'price' обязательно.")
                                    continue
                                price = float(row['price'])
                                if price < 0:
                                    errors.append(f"Запись {index + 1}: Цена не может быть отрицательной.")
                                    continue
                                if not row.get('category_path'):
                                    errors.append(f"Запись {index + 1}: Поле 'category_path' обязательно.")
                                    continue

                                category = self.get_or_create_category(row['category_path'])
                                if not category:
                                    errors.append(f"Запись {index + 1}: Не удалось создать категорию '{row['category_path']}'.")
                                    continue

                                photo_path = None
                                if row.get('photo_filename'):
                                    photo_path = os.path.join('product_photos', row['photo_filename'])
                                    if not default_storage.exists(photo_path):
                                        errors.append(f"Запись {index + 1}: Файл фотографии '{photo_path}' не найден.")
                                        continue

                                product_id = row.get('id')
                                if product_id:
                                    try:
                                        product = Product.objects.get(id=product_id)
                                    except Product.DoesNotExist:
                                        product = Product()
                                else:
                                    product = Product()

                                product.name = row['name']
                                product.description = row.get('description', '')
                                product.price = price
                                product.category = category
                                product.is_active = row.get('is_active', True)
                                if photo_path:
                                    product.photo = photo_path
                                product.save()

                                imported_count += 1

                            except Exception as e:
                                errors.append(f"Запись {index + 1}: Ошибка: {str(e)}")

                    if imported_count > 0:
                        messages.success(request, f"Импортировано {imported_count} товаров.")
                    if errors:
                        for error in errors:
                            messages.error(request, error)
                    else:
                        messages.success(request, "Импорт завершён без ошибок.")

                except Exception as e:
                    logger.error(f"Ошибка при импорте: {str(e)}")
                    messages.error(request, f"Ошибка при импорте: {str(e)}")

                return redirect('admin:shop_product_changelist')

            else:
                logger.warning("Форма невалидна")
                logger.warning(form.errors)
                messages.error(request, "Форма заполнена некорректно. Проверьте введённые данные.")
                return render(request, 'admin/shop/order/import_products.html', {'form': form})

        # Если это POST-запрос от админки (выбор действия)
        if request.method == "POST" and 'action' in request.POST:
            logger.info("Получен POST-запрос от админки для действия import_products")
            form = ProductImportForm()
            context = {
                'form': form,
                'action': 'import_products',
                'opts': self.model._meta,
                'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
            }
            return render(request, 'admin/shop/order/import_products.html', context)

        # Если это GET-запрос, просто отображаем форму
        form = ProductImportForm()
        return render(request, 'admin/shop/order/import_products.html', {'form': form})

    import_products.short_description = "Импортировать товары"

    def export_products(self, request, queryset):
        logger.info("Начало экспорта товаров")

        # Если это POST-запрос от формы экспорта
        if 'export_form_submit' in request.POST:
            logger.info("Получен POST-запрос от формы экспорта")
            form = ProductExportForm(request.POST)
            if form.is_valid():
                logger.info("Форма валидна")
                file_format = form.cleaned_data['file_format']
                selected_fields = form.cleaned_data['fields']
                category = form.cleaned_data['category']
                is_active = form.cleaned_data['is_active']
                date_from = form.cleaned_data['date_from']
                date_to = form.cleaned_data['date_to']

                logger.info(f"Параметры формы: file_format={file_format}, selected_fields={selected_fields}, "
                           f"category={category}, is_active={is_active}, date_from={date_from}, date_to={date_to}")

                # Используем queryset для экспорта выбранных товаров
                selected_ids = request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)
                if selected_ids:
                    products_to_export = Product.objects.filter(id__in=selected_ids)
                    logger.info(f"Выбрано {products_to_export.count()} товаров через selected_ids")
                else:
                    products_to_export = Product.objects.all()
                    logger.info(f"Всего товаров в базе: {products_to_export.count()}")
                    if category:
                        products_to_export = products_to_export.filter(category=category)
                        logger.info(f"После фильтра по категории: {products_to_export.count()}")
                    if is_active:
                        products_to_export = products_to_export.filter(is_active=(is_active == '1'))
                        logger.info(f"После фильтра по статусу: {products_to_export.count()}")
                    if date_from:
                        products_to_export = products_to_export.filter(created_at__gte=date_from)
                        logger.info(f"После фильтра по дате (с): {products_to_export.count()}")
                    if date_to:
                        products_to_export = products_to_export.filter(created_at__lte=date_to)
                        logger.info(f"После фильтра по дате (по): {products_to_export.count()}")

                # Проверяем, есть ли товары для экспорта
                if not products_to_export.exists():
                    logger.warning("Нет товаров для экспорта после применения фильтров")
                    messages.warning(request, "Нет товаров для экспорта. Проверьте фильтры или добавьте товары в базу.")
                    return redirect('admin:shop_product_changelist')

                # Формируем данные для экспорта
                logger.info("Формирование данных для экспорта")
                data = []
                for product in products_to_export:
                    row = {}
                    if 'id' in selected_fields:
                        row['id'] = product.id
                    if 'name' in selected_fields:
                        row['name'] = product.name
                    if 'description' in selected_fields:
                        row['description'] = product.description
                    if 'price' in selected_fields:
                        row['price'] = float(product.price)
                    if 'category_path' in selected_fields:
                        row['category_path'] = self.get_category_path(product.category)
                    if 'photo_filename' in selected_fields:
                        row['photo_filename'] = os.path.basename(product.photo.name) if product.photo else ''
                    if 'is_active' in selected_fields:
                        row['is_active'] = product.is_active
                    if 'created_at' in selected_fields:
                        row['created_at'] = product.created_at.isoformat()
                    data.append(row)

                logger.info(f"Сформировано {len(data)} записей для экспорта")

                # Генерируем файл и отправляем его пользователю для скачивания
                try:
                    if file_format == 'json':
                        logger.info("Генерация JSON-файла")
                        response = HttpResponse(content_type='application/json')
                        response['Content-Disposition'] = 'attachment; filename="products_export.json"'
                        response.write(json.dumps(data, ensure_ascii=False, indent=2))
                        logger.info("Файл JSON сгенерирован и отправлен")
                    elif file_format == 'csv':
                        logger.info("Генерация CSV-файла")
                        response = HttpResponse(content_type='text/csv')
                        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
                        writer = csv.DictWriter(response, fieldnames=selected_fields)
                        writer.writeheader()
                        for row in data:
                            writer.writerow(row)
                        logger.info("Файл CSV сгенерирован и отправлен")
                    elif file_format == 'xlsx':
                        logger.info("Генерация XLSX-файла")
                        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
                        workbook = openpyxl.Workbook()
                        sheet = workbook.active
                        sheet.append(selected_fields)
                        for row in data:
                            sheet.append([row.get(field, '') for field in selected_fields])
                        workbook.save(response)
                        logger.info("Файл Excel сгенерирован и отправлен")
                    else:
                        logger.error(f"Неизвестный формат файла: {file_format}")
                        messages.error(request, "Неизвестный формат файла.")
                        return redirect('admin:shop_product_changelist')

                    return response

                except Exception as e:
                    logger.error(f"Ошибка при генерации файла: {str(e)}")
                    messages.error(request, f"Ошибка при генерации файла: {str(e)}")
                    return redirect('admin:shop_product_changelist')

            else:
                logger.warning("Форма невалидна")
                logger.warning(form.errors)
                messages.error(request, "Форма заполнена некорректно. Проверьте введённые данные.")
                return render(request, 'admin/shop/order/export_products.html', {'form': form})

        # Если это POST-запрос от админки (выбор действия)
        if request.method == "POST" and 'action' in request.POST:
            logger.info("Получен POST-запрос от админки для действия export_products")
            # Проверяем, выбраны ли объекты
            selected_ids = request.POST.getlist(admin.helpers.ACTION_CHECKBOX_NAME)
            if not selected_ids:
                self.message_user(request, "Пожалуйста, выберите хотя бы один товар для экспорта.", level=messages.WARNING)
                return redirect('admin:shop_product_changelist')

            # Отображаем форму экспорта
            form = ProductExportForm()
            context = {
                'form': form,
                'queryset': Product.objects.filter(id__in=selected_ids),
                'action': 'export_products',
                'opts': self.model._meta,
                'action_checkbox_name': admin.helpers.ACTION_CHECKBOX_NAME,
            }
            return render(request, 'admin/shop/order/export_products.html', context)

        # Если это GET-запрос, просто отображаем форму
        form = ProductExportForm()
        return render(request, 'admin/shop/order/export_products.html', {'form': form})

    export_products.short_description = "Экспортировать товары"

    def parse_json(self, file):
        content = file.read().decode('utf-8')
        return json.loads(content)

    def csv_to_json(self, file):
        content = file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(content))
        return [dict(row) for row in reader]

    def excel_to_json(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for header, value in zip(headers, row):
                row_data[header] = value
            data.append(row_data)
        return data

    def get_or_create_category(self, category_path):
        parts = category_path.split('/')
        parent = None
        for part in parts:
            part = part.strip()
            if not part:
                continue
            category, created = Category.objects.get_or_create(
                name=part,
                parent=parent,
                defaults={'is_active': True}
            )
            parent = category
        return parent

    def get_category_path(self, category):
        if not category:
            return ''
        ancestors = category.get_ancestors(include_self=True)
        return '/'.join(ancestor.name for ancestor in ancestors)
